"""
Extract worklogs from Jira ZYN project filtered by ERP Activity.
Includes all issues linked to epics that match the filter.
"""

import requests
from requests.auth import HTTPBasicAuth
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import json
from datetime import datetime
from collections import defaultdict
import csv
from pathlib import Path
import os
import argparse
from tqdm import tqdm
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor, as_completed
import pickle
import hashlib
import time
import logging
from logging.handlers import RotatingFileHandler
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Version
VERSION = "1.0.0"

# Load environment variables from .env file
load_dotenv()

# Configuration - Load from environment variables with fallback to defaults
JIRA_URL = os.getenv('JIRA_URL', 'https://your-jira-instance.atlassian.net')
JIRA_EMAIL = os.getenv('JIRA_EMAIL', 'your-email@example.com')
JIRA_API_TOKEN = os.getenv('JIRA_API_TOKEN', 'your-api-token-here')
PROJECT_KEY = os.getenv('PROJECT_KEY', 'ZYN')
ERP_ACTIVITY_FILTER = os.getenv('ERP_ACTIVITY_FILTER', 'ProjectTask-00000007118797')
LOG_LEVEL = int(os.getenv('LOG_LEVEL', '1'))  # 1 = standard, 2 = debug (shows [DEBUG] messages)

def setup_logging(log_level=1, log_dir='logs'):
    """
    Setup logging with both console and file output.

    Args:
        log_level: 1 for INFO, 2 for DEBUG
        log_dir: Directory to store log files (default: 'logs')

    Returns:
        logger: Configured logger instance
    """
    # Create logs directory if it doesn't exist
    log_dir_path = Path(log_dir)
    log_dir_path.mkdir(exist_ok=True)

    # Generate log filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_filename = log_dir_path / f'jira_extractor_{timestamp}.log'

    # Set log level based on input
    level = logging.DEBUG if log_level >= 2 else logging.INFO

    # Create logger
    logger = logging.getLogger('JiraWorklogExtractor')
    logger.setLevel(level)

    # Clear any existing handlers
    logger.handlers.clear()

    # Create formatters
    detailed_formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    console_formatter = logging.Formatter('%(message)s')

    # File handler with rotation (max 10 files, 10MB each)
    file_handler = RotatingFileHandler(
        log_filename,
        maxBytes=10*1024*1024,  # 10MB
        backupCount=10,
        encoding='utf-8'
    )
    file_handler.setLevel(level)
    file_handler.setFormatter(detailed_formatter)
    logger.addHandler(file_handler)

    # Console handler (INFO and above only for clean output)
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(console_formatter)
    logger.addHandler(console_handler)

    return logger

class JiraWorklogExtractor:
    def __init__(self, jira_url, api_token, project_key=None, erp_activity_filter=None, log_level=1, cache_dir='.cache', use_cache=True, cache_ttl=3600, skip_validation=False, logger=None):
        self.jira_url = jira_url.rstrip('/')
        self.api_token = api_token
        self.project_key = project_key or PROJECT_KEY
        self.erp_activity_filter = erp_activity_filter or ERP_ACTIVITY_FILTER
        self.log_level = log_level
        self.use_cache = use_cache
        self.cache_ttl = cache_ttl  # Cache time-to-live in seconds (default: 1 hour)
        self.logger = logger or logging.getLogger('JiraWorklogExtractor')
        self.headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {api_token}"
        }

        # Set up HTTP session with retry logic
        self.session = self._create_session_with_retries()

        # Set up cache directory
        self.cache_dir = Path(cache_dir)
        if self.use_cache:
            self.cache_dir.mkdir(exist_ok=True)
            self.logger.debug(f"Cache enabled: {self.cache_dir.absolute()} (TTL: {cache_ttl}s)")

        # Validate configuration if not skipped
        if not skip_validation:
            self.validate_configuration()

    def _create_session_with_retries(self):
        """Create HTTP session with automatic retry logic"""
        session = requests.Session()
        session.headers.update(self.headers)

        # Configure retry strategy
        retry_strategy = Retry(
            total=3,  # Maximum number of retries
            backoff_factor=1,  # Wait 1, 2, 4 seconds between retries
            status_forcelist=[429, 500, 502, 503, 504],  # Retry on these HTTP status codes
            allowed_methods=["GET", "POST"],  # Only retry safe methods
            raise_on_status=False  # Don't raise exception, let us handle it
        )

        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("http://", adapter)
        session.mount("https://", adapter)

        return session

    def validate_configuration(self):
        """
        Validate Jira configuration settings.

        Checks:
        1. Jira URL is reachable
        2. API token is valid (authentication works)
        3. Project exists
        4. Custom fields exist (customfield_10014, customfield_11440, customfield_10076)
        5. ERP Activity filter field exists

        Raises:
            Exception: If any validation check fails with descriptive error message

        Returns:
            bool: True if all validations pass
        """
        print(f"\n{'='*70}")
        print("VALIDATING CONFIGURATION")
        print(f"{'='*70}")

        # 1. Check if Jira URL is reachable
        print("\n[1/5] Checking Jira URL accessibility...")
        try:
            response = self.session.get(self.jira_url, timeout=10)
            if response.status_code >= 400:
                raise Exception(f"Jira URL returned status code {response.status_code}")
            print(f"  [OK] Jira URL is reachable: {self.jira_url}")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Failed to reach Jira URL '{self.jira_url}': {e}")

        # 2. Verify API token is valid
        print("\n[2/5] Verifying API token authentication...")
        try:
            url = f"{self.jira_url}/rest/api/2/myself"
            response = self.session.get(url, timeout=10)

            if response.status_code == 401:
                raise Exception("API token is invalid or expired. Please check your credentials.")
            elif response.status_code == 403:
                raise Exception("API token does not have sufficient permissions.")
            elif response.status_code >= 400:
                raise Exception(f"Authentication failed with status code {response.status_code}: {response.text[:200]}")

            user_data = response.json()
            user_name = user_data.get('displayName', 'Unknown')
            user_email = user_data.get('emailAddress', 'Unknown')
            print(f"  [OK] API token is valid")
            print(f"    Authenticated as: {user_name} ({user_email})")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Failed to verify API token: {e}")

        # 3. Confirm the project exists
        print(f"\n[3/5] Verifying project '{self.project_key}' exists...")
        try:
            url = f"{self.jira_url}/rest/api/2/project/{self.project_key}"
            response = self.session.get(url, timeout=10)

            if response.status_code == 404:
                raise Exception(f"Project '{self.project_key}' does not exist or you don't have access to it.")
            elif response.status_code >= 400:
                raise Exception(f"Failed to retrieve project with status code {response.status_code}: {response.text[:200]}")

            project_data = response.json()
            project_name = project_data.get('name', 'Unknown')
            print(f"  [OK] Project exists: {self.project_key} - {project_name}")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Failed to verify project: {e}")

        # 4. Test that custom fields exist
        print("\n[4/5] Verifying custom fields exist...")
        custom_fields = {
            'customfield_10014': 'Epic Link',
            'customfield_11440': 'Product Item',
            'customfield_10076': 'Team Name'
        }

        try:
            # Get a sample issue from the project to check custom fields
            url = f"{self.jira_url}/rest/api/2/search"
            params = {
                "jql": f"project = {self.project_key}",
                "maxResults": 1,
                "fields": ",".join(custom_fields.keys())
            }
            response = self.session.get(url, params=params, timeout=10)

            if response.status_code >= 400:
                raise Exception(f"Failed to query issues with status code {response.status_code}: {response.text[:200]}")

            data = response.json()

            if data.get('total', 0) == 0:
                print(f"  [!] Warning: No issues found in project '{self.project_key}' to validate custom fields")
                print("    Custom fields will be validated when issues are processed")
            else:
                # Check if custom fields are present in the response
                issue = data['issues'][0]
                fields = issue.get('fields', {})

                missing_fields = []
                for field_id, field_name in custom_fields.items():
                    if field_id in fields:
                        print(f"  [OK] Custom field exists: {field_name} ({field_id})")
                    else:
                        missing_fields.append(f"{field_name} ({field_id})")

                if missing_fields:
                    print(f"  [!] Warning: The following custom fields were not found:")
                    for field in missing_fields:
                        print(f"    - {field}")
                    print("    These fields may not be configured for this project or issue type")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Failed to verify custom fields: {e}")

        # 5. Validate the ERP Activity filter field exists
        print("\n[5/5] Validating ERP Activity field...")
        try:
            # Try to query with the ERP Activity field to see if it exists
            url = f"{self.jira_url}/rest/api/2/search"
            jql = f'project = {self.project_key} AND "ERP Activity" IS NOT EMPTY'
            params = {
                "jql": jql,
                "maxResults": 1
            }
            response = self.session.get(url, params=params, timeout=10)

            if response.status_code >= 400:
                error_message = response.text
                if "does not exist" in error_message or "not exist" in error_message:
                    raise Exception(f"ERP Activity field does not exist in project '{self.project_key}'")
                else:
                    raise Exception(f"Failed to validate ERP Activity field with status code {response.status_code}: {response.text[:200]}")

            data = response.json()
            issue_count = data.get('total', 0)
            print(f"  [OK] ERP Activity field exists")
            print(f"    Found {issue_count} issue(s) with ERP Activity field populated")

            # Test the specific filter value
            jql_with_filter = f'project = {self.project_key} AND "ERP Activity" ~ "{self.erp_activity_filter}"'
            params['jql'] = jql_with_filter
            response = self.session.get(url, params=params, timeout=10)

            if response.status_code >= 400:
                raise Exception(f"Failed to query with ERP Activity filter with status code {response.status_code}: {response.text[:200]}")

            data = response.json()
            filter_count = data.get('total', 0)

            if filter_count == 0:
                print(f"  [!] Warning: No issues found matching ERP Activity filter '{self.erp_activity_filter}'")
                print("    Please verify the filter value is correct")
            else:
                print(f"    Found {filter_count} issue(s) matching filter '{self.erp_activity_filter}'")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Failed to validate ERP Activity field: {e}")

        print(f"\n{'='*70}")
        print("CONFIGURATION VALIDATION COMPLETE")
        print(f"{'='*70}\n")

        return True

    def _get_cache_key(self, prefix, identifier):
        """Generate a cache key from prefix and identifier"""
        # Create hash of identifier to handle long strings
        hash_obj = hashlib.md5(str(identifier).encode())
        return f"{prefix}_{hash_obj.hexdigest()}.pkl"

    def _get_from_cache(self, cache_key):
        """Retrieve data from cache if valid"""
        if not self.use_cache:
            return None

        cache_file = self.cache_dir / cache_key
        if not cache_file.exists():
            return None

        try:
            # Check if cache is expired
            file_age = time.time() - cache_file.stat().st_mtime
            if file_age > self.cache_ttl:
                self.logger.debug(f"Cache expired: {cache_key} (age: {file_age:.0f}s)")
                cache_file.unlink()  # Delete expired cache
                return None

            with open(cache_file, 'rb') as f:
                data = pickle.load(f)
                self.logger.debug(f"Cache hit: {cache_key}")
                return data
        except Exception as e:
            self.logger.warning(f"Failed to read cache {cache_key}: {e}")
            return None

    def _save_to_cache(self, cache_key, data):
        """Save data to cache"""
        if not self.use_cache:
            return

        cache_file = self.cache_dir / cache_key
        try:
            with open(cache_file, 'wb') as f:
                pickle.dump(data, f)
                self.logger.debug(f"Cache saved: {cache_key}")
        except Exception as e:
            self.logger.warning(f"Failed to save cache {cache_key}: {e}")

    def search_issues(self, jql, fields=None, max_results=100):
        """Search for issues using JQL"""
        url = f"{self.jira_url}/rest/api/2/search"
        all_issues = []
        start_at = 0

        if fields is None:
            fields = ["summary", "issuetype", "status", "issuelinks"]

        while True:
            params = {
                "jql": jql,
                "startAt": start_at,
                "maxResults": max_results,
                "fields": ",".join(fields)
            }

            if self.log_level >= 2:
                print(f"\n[DEBUG] API Call: GET {url}")
                print(f"[DEBUG] JQL: {jql}")
                print(f"[DEBUG] Parameters: {params}")
                print(f"[DEBUG] Headers: {self.headers}")

            response = self.session.get(url, params=params)

            if self.log_level >= 2:
                print(f"[DEBUG] Response Status: {response.status_code}")

            if response.status_code != 200:
                if self.log_level >= 2:
                    print(f"[DEBUG] Response Body: {response.text[:500]}")

            response.raise_for_status()
            data = response.json()

            batch_size = len(data['issues'])
            all_issues.extend(data['issues'])

            if self.log_level >= 2:
                print(f"[DEBUG] Retrieved {batch_size} issues (total so far: {len(all_issues)}/{data['total']})")
            else:
                print(f"  Fetching issues: {len(all_issues)}/{data['total']}")

            if start_at + max_results >= data['total']:
                break
            start_at += max_results

        return all_issues

    def get_epic_issues(self, epic_key):
        """Get all issues in an epic"""
        print(f"  Searching for issues in epic {epic_key}...")
        jql = f'project = {self.project_key} AND "Epic Link" = {epic_key}'
        return self.search_issues(jql)

    def get_linked_issues(self, issue_key):
        """Get all issues linked to a given issue"""
        url = f"{self.jira_url}/rest/api/2/issue/{issue_key}"
        params = {"fields": "issuelinks"}

        if self.log_level >= 2:
            print(f"\n[DEBUG] Getting linked issues for {issue_key}")
            print(f"[DEBUG] API Call: GET {url}")

        response = self.session.get(url, params=params)

        if self.log_level >= 2:
            print(f"[DEBUG] Response Status: {response.status_code}")

        if response.status_code != 200:
            if self.log_level >= 2:
                print(f"[DEBUG] Response Body: {response.text[:500]}")

        response.raise_for_status()
        data = response.json()

        linked_issues = []
        issue_links = data.get('fields', {}).get('issuelinks', [])

        for link in issue_links:
            if 'outwardIssue' in link:
                linked_issues.append(link['outwardIssue']['key'])
            if 'inwardIssue' in link:
                linked_issues.append(link['inwardIssue']['key'])

        if linked_issues and LOG_LEVEL >= 2:
            print(f"[DEBUG] Found {len(linked_issues)} linked issues: {', '.join(linked_issues)}")

        return linked_issues

    def get_worklogs(self, issue_key):
        """Get all worklogs for an issue (with caching)"""
        # Check cache first
        cache_key = self._get_cache_key("worklogs", issue_key)
        cached_data = self._get_from_cache(cache_key)
        if cached_data is not None:
            return cached_data

        url = f"{self.jira_url}/rest/api/2/issue/{issue_key}/worklog"

        if self.log_level >= 2:
            print(f"[DEBUG] Getting worklogs for {issue_key}")

        response = self.session.get(url)

        if self.log_level >= 2:
            print(f"[DEBUG] Response Status: {response.status_code}")

        if response.status_code != 200:
            if self.log_level >= 2:
                print(f"[DEBUG] Response Body: {response.text[:500]}")

        response.raise_for_status()
        data = response.json()

        worklogs = data.get('worklogs', [])
        if worklogs and LOG_LEVEL >= 2:
            print(f"[DEBUG] Found {len(worklogs)} worklog(s) for {issue_key}")

        # Save to cache
        self._save_to_cache(cache_key, worklogs)

        return worklogs

    def find_issues_by_erp_activity(self):
        """Find all issues with ERP Activity filter"""
        jql = f'project = {self.project_key} AND "ERP Activity" ~ "{self.erp_activity_filter}"'
        return self.search_issues(jql)

    def get_subtasks(self, issue_key):
        """Get all subtasks for an issue"""
        url = f"{self.jira_url}/rest/api/2/issue/{issue_key}"
        params = {"fields": "subtasks"}

        if self.log_level >= 2:
            print(f"\n[DEBUG] Getting subtasks for {issue_key}")

        response = self.session.get(url, params=params)
        response.raise_for_status()
        data = response.json()

        subtasks = data.get('fields', {}).get('subtasks', [])
        subtask_keys = [subtask['key'] for subtask in subtasks]

        if subtask_keys and LOG_LEVEL >= 2:
            print(f"[DEBUG] Found {len(subtask_keys)} subtask(s): {', '.join(subtask_keys)}")

        return subtask_keys

    def collect_all_related_issues(self):
        """
        Collect all issues related to the ERP Activity:
        1. Direct matches
        2. Issues linked to epics
        3. Issues linked to any matched issue
        4. Sub-tasks of all collected issues
        """
        print(f"\n{'='*70}")
        print(f"STARTING ISSUE COLLECTION")
        print(f"{'='*70}")
        print(f'Searching for issues with "ERP Activity" ~ "{self.erp_activity_filter}"...')

        # Step 1: Find all issues matching the ERP Activity
        direct_issues = self.find_issues_by_erp_activity()
        print(f"\n[STEP 1 COMPLETE] Found {len(direct_issues)} direct matches")

        all_issue_keys = set()
        epic_keys = []

        # Collect direct issue keys and identify epics
        print("\n[STEP 2] Analyzing direct matches for epics...")
        for issue in direct_issues:
            issue_key = issue['key']
            all_issue_keys.add(issue_key)
            print(f"  - {issue_key}")

            issue_type = issue['fields'].get('issuetype', {}).get('name', '')
            if issue_type == 'Epic':
                epic_keys.append(issue_key)

        print(f"\n[STEP 2 COMPLETE] Found {len(epic_keys)} epics: {epic_keys}")

        # Step 2: Get all issues linked to epics
        if epic_keys:
            print(f"\n[STEP 3] Getting issues from epics...")
            for epic_key in epic_keys:
                print(f"\n  Processing epic {epic_key}...")
                epic_issues = self.get_epic_issues(epic_key)
                before_count = len(all_issue_keys)
                for issue in epic_issues:
                    all_issue_keys.add(issue['key'])
                after_count = len(all_issue_keys)
                print(f"  [EPIC COMPLETE] Added {after_count - before_count} new issues from epic {epic_key}")
            print(f"\n[STEP 3 COMPLETE] Total issues after epic expansion: {len(all_issue_keys)}")
        else:
            print(f"\n[STEP 3 SKIPPED] No epics found")

        # Step 3: Get all linked issues
        print(f"\n[STEP 4] Finding linked issues (checking {len(all_issue_keys)} issues)...")
        initial_keys = list(all_issue_keys)
        linked_count = 0
        for issue_key in tqdm(initial_keys, desc="Finding linked issues", unit="issue"):
            linked = self.get_linked_issues(issue_key)
            for linked_key in linked:
                if linked_key.startswith(self.project_key):
                    if linked_key not in all_issue_keys:
                        all_issue_keys.add(linked_key)
                        linked_count += 1

        print(f"[STEP 4 COMPLETE] Added {linked_count} new linked issues")

        # Step 4: Get all subtasks
        print(f"\n[STEP 5] Finding subtasks (checking {len(all_issue_keys)} issues)...")
        initial_keys_for_subtasks = list(all_issue_keys)
        subtask_count = 0
        for issue_key in tqdm(initial_keys_for_subtasks, desc="Finding subtasks", unit="issue"):
            subtasks = self.get_subtasks(issue_key)
            for subtask_key in subtasks:
                if subtask_key not in all_issue_keys:
                    all_issue_keys.add(subtask_key)
                    subtask_count += 1

        print(f"[STEP 5 COMPLETE] Added {subtask_count} new subtasks")
        print(f"\n{'='*70}")
        print(f"ISSUE COLLECTION COMPLETE: {len(all_issue_keys)} total issues")
        print(f"{'='*70}\n")
        return sorted(all_issue_keys)

    def get_issue_metadata(self, issue_key):
        """Get issue type, epic link, summary, components, labels, and team for an issue (with caching)"""
        # Check cache first
        cache_key = self._get_cache_key("metadata", issue_key)
        cached_data = self._get_from_cache(cache_key)
        if cached_data is not None:
            return cached_data

        url = f"{self.jira_url}/rest/api/2/issue/{issue_key}"
        # Fetch additional fields: components, labels, customfield_10014 (Epic Link), customfield_11440 (Product Item), customfield_10076 (Team Name)
        params = {"fields": "issuetype,customfield_10014,summary,components,labels,customfield_11440,customfield_10076"}

        response = self.session.get(url, params=params)
        response.raise_for_status()
        data = response.json()

        fields = data.get('fields', {})

        # Extract components (array of objects with 'name')
        components = fields.get('components', [])
        component_names = [comp.get('name', '') for comp in components if comp.get('name')]

        # Extract labels (array of strings)
        labels = fields.get('labels', [])

        # Extract Product Item (custom field)
        product_item = fields.get('customfield_11440', '')
        if isinstance(product_item, dict):
            product_item = product_item.get('value', '') or product_item.get('name', '')

        # Extract Team Name (custom field - prioritize 'name' property)
        team = fields.get('customfield_10076', '')
        if isinstance(team, dict):
            team = team.get('name', '') or team.get('value', '')

        metadata = {
            'issue_type': fields.get('issuetype', {}).get('name', 'Unknown'),
            'epic_link': fields.get('customfield_10014', ''),
            'summary': fields.get('summary', ''),
            'components': component_names,
            'labels': labels,
            'product_item': product_item or 'None',
            'team': team or 'None'
        }

        # Save to cache
        self._save_to_cache(cache_key, metadata)

        return metadata

    def _extract_issue_worklogs(self, issue_key, issue_metadata_cache):
        """Extract worklogs from a single issue (helper for parallel execution)"""
        result_worklogs = []
        try:
            # Get issue metadata
            if issue_key not in issue_metadata_cache:
                issue_metadata_cache[issue_key] = self.get_issue_metadata(issue_key)

            metadata = issue_metadata_cache[issue_key]

            worklogs = self.get_worklogs(issue_key)
            for worklog in worklogs:
                # Handle comment field - can be string, dict, or None
                comment = ''
                if worklog.get('comment'):
                    comment_field = worklog['comment']
                    if isinstance(comment_field, str):
                        comment = comment_field
                    elif isinstance(comment_field, dict):
                        # Jira Atlassian Document Format
                        try:
                            content = comment_field.get('content', [])
                            if content and len(content) > 0:
                                first_block = content[0]
                                if isinstance(first_block, dict):
                                    block_content = first_block.get('content', [])
                                    if block_content and len(block_content) > 0:
                                        text_node = block_content[0]
                                        if isinstance(text_node, dict):
                                            comment = text_node.get('text', '')
                        except (KeyError, IndexError, AttributeError):
                            comment = str(comment_field)

                result_worklogs.append({
                    'issue_key': issue_key,
                    'issue_type': metadata['issue_type'],
                    'epic_link': metadata['epic_link'],
                    'summary': metadata['summary'],
                    'components': metadata.get('components', []),
                    'labels': metadata.get('labels', []),
                    'product_item': metadata.get('product_item', 'None'),
                    'team': metadata.get('team', 'None'),
                    'worklog_id': worklog['id'],
                    'author': worklog['author'].get('displayName', 'Unknown'),
                    'author_email': worklog['author'].get('emailAddress', ''),
                    'time_spent': worklog.get('timeSpent', ''),
                    'time_spent_seconds': worklog.get('timeSpentSeconds', 0),
                    'started': worklog.get('started', ''),
                    'comment': comment
                })
        except Exception as e:
            if self.log_level >= 1:
                print(f"  Error extracting worklogs from {issue_key}: {e}")

        return result_worklogs

    def extract_worklogs(self, issue_keys, max_workers=10):
        """Extract worklogs from all issues using parallel requests"""
        all_worklogs = []
        issue_metadata_cache = {}

        print(f"\nExtracting worklogs from {len(issue_keys)} issues (using {max_workers} parallel workers)...")

        # Use ThreadPoolExecutor for parallel API calls
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all tasks
            future_to_key = {
                executor.submit(self._extract_issue_worklogs, key, issue_metadata_cache): key
                for key in issue_keys
            }

            # Process results as they complete with progress bar
            for future in tqdm(as_completed(future_to_key),
                             total=len(issue_keys),
                             desc="Extracting worklogs",
                             unit="issue"):
                try:
                    worklogs = future.result()
                    all_worklogs.extend(worklogs)
                except Exception as e:
                    issue_key = future_to_key[future]
                    if self.log_level >= 1:
                        print(f"  Error processing {issue_key}: {e}")

        print(f"\nTotal worklogs extracted: {len(all_worklogs)}")
        return all_worklogs

    def export_to_csv(self, worklogs, filename='worklogs_export.csv'):
        """Export worklogs to CSV"""
        if not worklogs:
            print("No worklogs to export")
            return

        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['issue_key', 'summary', 'issue_type', 'epic_link', 'author', 'author_email', 'time_spent',
                         'time_spent_hours', 'started', 'comment']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()
            for worklog in worklogs:
                writer.writerow({
                    'issue_key': worklog['issue_key'],
                    'summary': worklog['summary'],
                    'issue_type': worklog['issue_type'],
                    'epic_link': worklog['epic_link'],
                    'author': worklog['author'],
                    'author_email': worklog['author_email'],
                    'time_spent': worklog['time_spent'],
                    'time_spent_hours': round(worklog['time_spent_seconds'] / 3600, 2),
                    'started': worklog['started'],
                    'comment': worklog['comment']
                })

        print(f"Worklogs exported to {filename}")

    def export_to_excel(self, worklogs, filename='worklogs_export.xlsx'):
        """
        Export worklogs to Excel with multiple sheets and formatting.

        Sheets included:
        - Summary: Overall statistics
        - Hours by Year/Month: Time aggregated by month
        - Hours by Product Item: Time by product
        - Hours by Component: Time by component
        - Hours by Label: Time by label
        - Hours by Team: Time by team
        - Hours by Author: Time by contributor
        - Hours by Issue: Time by issue
        - All Worklog Entries: Raw worklog data
        """
        if not worklogs:
            print("No worklogs to export")
            return

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Calculate aggregations
        by_author = defaultdict(lambda: {'hours': 0, 'entries': 0})
        by_issue = defaultdict(lambda: {'hours': 0, 'entries': 0, 'summary': '', 'issue_type': ''})
        by_year_month = defaultdict(lambda: {'hours': 0, 'entries': 0})
        by_product_item = defaultdict(lambda: {'hours': 0, 'entries': 0, 'issues': set()})
        by_component = defaultdict(lambda: {'hours': 0, 'entries': 0, 'issues': set()})
        by_label = defaultdict(lambda: {'hours': 0, 'entries': 0, 'issues': set()})
        by_team = defaultdict(lambda: {'hours': 0, 'entries': 0, 'issues': set()})
        total_seconds = 0

        for worklog in worklogs:
            author = worklog['author']
            issue_key = worklog['issue_key']
            seconds = worklog['time_spent_seconds']

            by_author[author]['hours'] += seconds / 3600
            by_author[author]['entries'] += 1

            by_issue[issue_key]['hours'] += seconds / 3600
            by_issue[issue_key]['entries'] += 1
            by_issue[issue_key]['summary'] = worklog['summary']
            by_issue[issue_key]['issue_type'] = worklog['issue_type']

            # Year/Month
            if worklog['started']:
                year_month = worklog['started'][:7]
                by_year_month[year_month]['hours'] += seconds / 3600
                by_year_month[year_month]['entries'] += 1

            # Product Item
            product_item = worklog.get('product_item', 'None')
            by_product_item[product_item]['hours'] += seconds / 3600
            by_product_item[product_item]['entries'] += 1
            by_product_item[product_item]['issues'].add(issue_key)

            # Component
            components = worklog.get('components', [])
            component_key = ', '.join(sorted(components)) if components else 'None'
            by_component[component_key]['hours'] += seconds / 3600
            by_component[component_key]['entries'] += 1
            by_component[component_key]['issues'].add(issue_key)

            # Label
            labels = worklog.get('labels', [])
            label_key = ', '.join(sorted(labels)) if labels else 'None'
            by_label[label_key]['hours'] += seconds / 3600
            by_label[label_key]['entries'] += 1
            by_label[label_key]['issues'].add(issue_key)

            # Team
            team = worklog.get('team', 'None')
            by_team[team]['hours'] += seconds / 3600
            by_team[team]['entries'] += 1
            by_team[team]['issues'].add(issue_key)

            total_seconds += seconds

        total_hours = total_seconds / 3600

        # Helper function to format headers
        def format_header_row(ws, row_num, headers):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

        # Helper function to auto-size columns
        def auto_size_columns(ws):
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary")
        summary_data = [
            ["Metric", "Value"],
            ["Total Hours", round(total_hours, 2)],
            ["Total Entries", len(worklogs)],
            ["Contributors", len(by_author)],
            ["Issues", len(by_issue)],
            ["Product Items", len(by_product_item)],
            ["Components", len(by_component)],
            ["Labels", len(by_label)],
            ["Teams", len(by_team)]
        ]
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                if row_num == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
        auto_size_columns(ws_summary)

        # Sheet 2: Hours by Year/Month
        ws_year_month = wb.create_sheet("Hours by Year-Month")
        format_header_row(ws_year_month, 1, ["Year-Month", "Hours", "Entries"])
        for row_num, (year_month, data) in enumerate(sorted(by_year_month.items()), 2):
            ws_year_month.cell(row=row_num, column=1).value = year_month
            ws_year_month.cell(row=row_num, column=2).value = round(data['hours'], 2)
            ws_year_month.cell(row=row_num, column=3).value = data['entries']
            for col in range(1, 4):
                ws_year_month.cell(row=row_num, column=col).border = border
        auto_size_columns(ws_year_month)

        # Sheet 3: Hours by Product Item
        ws_product = wb.create_sheet("Hours by Product Item")
        format_header_row(ws_product, 1, ["Product Item", "Hours", "Entries", "Issues"])
        for row_num, (product, data) in enumerate(sorted(by_product_item.items(), key=lambda x: x[1]['hours'], reverse=True), 2):
            ws_product.cell(row=row_num, column=1).value = product
            ws_product.cell(row=row_num, column=2).value = round(data['hours'], 2)
            ws_product.cell(row=row_num, column=3).value = data['entries']
            ws_product.cell(row=row_num, column=4).value = len(data['issues'])
            for col in range(1, 5):
                ws_product.cell(row=row_num, column=col).border = border
        auto_size_columns(ws_product)

        # Sheet 4: Hours by Component
        ws_component = wb.create_sheet("Hours by Component")
        format_header_row(ws_component, 1, ["Component", "Hours", "Entries", "Issues"])
        for row_num, (component, data) in enumerate(sorted(by_component.items(), key=lambda x: x[1]['hours'], reverse=True), 2):
            ws_component.cell(row=row_num, column=1).value = component
            ws_component.cell(row=row_num, column=2).value = round(data['hours'], 2)
            ws_component.cell(row=row_num, column=3).value = data['entries']
            ws_component.cell(row=row_num, column=4).value = len(data['issues'])
            for col in range(1, 5):
                ws_component.cell(row=row_num, column=col).border = border
        auto_size_columns(ws_component)

        # Sheet 5: Hours by Label
        ws_label = wb.create_sheet("Hours by Label")
        format_header_row(ws_label, 1, ["Label", "Hours", "Entries", "Issues"])
        for row_num, (label, data) in enumerate(sorted(by_label.items(), key=lambda x: x[1]['hours'], reverse=True), 2):
            ws_label.cell(row=row_num, column=1).value = label
            ws_label.cell(row=row_num, column=2).value = round(data['hours'], 2)
            ws_label.cell(row=row_num, column=3).value = data['entries']
            ws_label.cell(row=row_num, column=4).value = len(data['issues'])
            for col in range(1, 5):
                ws_label.cell(row=row_num, column=col).border = border
        auto_size_columns(ws_label)

        # Sheet 6: Hours by Team
        ws_team = wb.create_sheet("Hours by Team")
        format_header_row(ws_team, 1, ["Team", "Hours", "Entries", "Issues"])
        for row_num, (team, data) in enumerate(sorted(by_team.items(), key=lambda x: x[1]['hours'], reverse=True), 2):
            ws_team.cell(row=row_num, column=1).value = team
            ws_team.cell(row=row_num, column=2).value = round(data['hours'], 2)
            ws_team.cell(row=row_num, column=3).value = data['entries']
            ws_team.cell(row=row_num, column=4).value = len(data['issues'])
            for col in range(1, 5):
                ws_team.cell(row=row_num, column=col).border = border
        auto_size_columns(ws_team)

        # Sheet 7: Hours by Author
        ws_author = wb.create_sheet("Hours by Author")
        format_header_row(ws_author, 1, ["Author", "Hours", "Entries"])
        for row_num, (author, data) in enumerate(sorted(by_author.items(), key=lambda x: x[1]['hours'], reverse=True), 2):
            ws_author.cell(row=row_num, column=1).value = author
            ws_author.cell(row=row_num, column=2).value = round(data['hours'], 2)
            ws_author.cell(row=row_num, column=3).value = data['entries']
            for col in range(1, 4):
                ws_author.cell(row=row_num, column=col).border = border
        auto_size_columns(ws_author)

        # Sheet 8: Hours by Issue
        ws_issue = wb.create_sheet("Hours by Issue")
        format_header_row(ws_issue, 1, ["Issue Key", "Summary", "Issue Type", "Hours", "Entries"])
        for row_num, (issue, data) in enumerate(sorted(by_issue.items(), key=lambda x: x[1]['hours'], reverse=True), 2):
            ws_issue.cell(row=row_num, column=1).value = issue
            ws_issue.cell(row=row_num, column=2).value = data['summary']
            ws_issue.cell(row=row_num, column=3).value = data['issue_type']
            ws_issue.cell(row=row_num, column=4).value = round(data['hours'], 2)
            ws_issue.cell(row=row_num, column=5).value = data['entries']
            for col in range(1, 6):
                ws_issue.cell(row=row_num, column=col).border = border
        auto_size_columns(ws_issue)

        # Sheet 9: All Worklog Entries
        ws_entries = wb.create_sheet("All Worklog Entries")
        headers = ['Issue Key', 'Summary', 'Issue Type', 'Epic Link', 'Author', 'Author Email',
                   'Time Spent', 'Hours', 'Started', 'Product Item', 'Team', 'Comment']
        format_header_row(ws_entries, 1, headers)

        for row_num, worklog in enumerate(worklogs, 2):
            ws_entries.cell(row=row_num, column=1).value = worklog['issue_key']
            ws_entries.cell(row=row_num, column=2).value = worklog['summary']
            ws_entries.cell(row=row_num, column=3).value = worklog['issue_type']
            ws_entries.cell(row=row_num, column=4).value = worklog['epic_link']
            ws_entries.cell(row=row_num, column=5).value = worklog['author']
            ws_entries.cell(row=row_num, column=6).value = worklog['author_email']
            ws_entries.cell(row=row_num, column=7).value = worklog['time_spent']
            ws_entries.cell(row=row_num, column=8).value = round(worklog['time_spent_seconds'] / 3600, 2)
            ws_entries.cell(row=row_num, column=9).value = worklog['started']
            ws_entries.cell(row=row_num, column=10).value = worklog.get('product_item', '')
            ws_entries.cell(row=row_num, column=11).value = worklog.get('team', '')
            ws_entries.cell(row=row_num, column=12).value = worklog['comment']
            for col in range(1, 13):
                ws_entries.cell(row=row_num, column=col).border = border
        auto_size_columns(ws_entries)

        # Save workbook
        wb.save(filename)
        print(f"Worklogs exported to Excel: {filename}")

    def generate_summary(self, worklogs):
        """Generate summary statistics"""
        if not worklogs:
            print("No worklogs to summarize")
            return

        # Summary by author
        by_author = defaultdict(lambda: {'hours': 0, 'entries': 0})
        total_seconds = 0

        for worklog in worklogs:
            author = worklog['author']
            seconds = worklog['time_spent_seconds']
            by_author[author]['hours'] += seconds / 3600
            by_author[author]['entries'] += 1
            total_seconds += seconds

        print("\n" + "="*60)
        print("SUMMARY REPORT")
        print("="*60)
        print(f"Total hours logged: {total_seconds / 3600:.2f}")
        print(f"Total worklog entries: {len(worklogs)}")
        print("\nBreakdown by author:")
        print("-"*60)

        for author, stats in sorted(by_author.items(), key=lambda x: x[1]['hours'], reverse=True):
            print(f"{author:30s} {stats['hours']:8.2f} hours  ({stats['entries']} entries)")

    def export_to_html(self, worklogs, filename='worklogs_report.html', timing_info=None):
        """Export worklogs to an interactive HTML report"""
        if not worklogs:
            print("No worklogs to export")
            return

        # Helper function to create Jira issue link
        def make_issue_link(issue_key):
            jira_url = f"{self.jira_url}/browse/{issue_key}"
            return f'<a href="{jira_url}" target="_blank" style="color: #0052CC; text-decoration: none; font-weight: bold;">{issue_key}</a>'

        # Calculate summary statistics
        by_author = defaultdict(lambda: {'hours': 0, 'entries': 0, 'worklogs': []})
        by_issue = defaultdict(lambda: {'hours': 0, 'entries': 0, 'authors': set(), 'issue_type': '', 'epic_link': '', 'summary': ''})
        by_year_month = defaultdict(lambda: {'hours': 0, 'entries': 0})
        by_product_item = defaultdict(lambda: {'hours': 0, 'entries': 0, 'issues': set()})
        by_component = defaultdict(lambda: {'hours': 0, 'entries': 0, 'issues': set()})
        by_label = defaultdict(lambda: {'hours': 0, 'entries': 0, 'issues': set()})
        by_team = defaultdict(lambda: {'hours': 0, 'entries': 0, 'issues': set()})
        total_seconds = 0

        for worklog in worklogs:
            author = worklog['author']
            issue_key = worklog['issue_key']
            seconds = worklog['time_spent_seconds']

            by_author[author]['hours'] += seconds / 3600
            by_author[author]['entries'] += 1
            by_author[author]['worklogs'].append(worklog)

            by_issue[issue_key]['hours'] += seconds / 3600
            by_issue[issue_key]['entries'] += 1
            by_issue[issue_key]['authors'].add(author)
            by_issue[issue_key]['issue_type'] = worklog['issue_type']
            by_issue[issue_key]['epic_link'] = worklog['epic_link']
            by_issue[issue_key]['summary'] = worklog['summary']
            by_issue[issue_key]['components'] = worklog.get('components', [])
            by_issue[issue_key]['labels'] = worklog.get('labels', [])
            by_issue[issue_key]['product_item'] = worklog.get('product_item', 'None')
            by_issue[issue_key]['team'] = worklog.get('team', 'None')

            # Extract year and month from worklog started date
            if worklog['started']:
                # started format: 2024-03-15T10:30:00.000+0000
                date_part = worklog['started'][:10]  # Get YYYY-MM-DD
                year_month = date_part[:7]  # Get YYYY-MM
                by_year_month[year_month]['hours'] += seconds / 3600
                by_year_month[year_month]['entries'] += 1

            # Aggregate by Product Item
            product_item = worklog.get('product_item', 'None')
            by_product_item[product_item]['hours'] += seconds / 3600
            by_product_item[product_item]['entries'] += 1
            by_product_item[product_item]['issues'].add(issue_key)

            # Aggregate by Component (keep multiple components together, don't split)
            components = worklog.get('components', [])
            if components:
                # Join multiple components with comma to keep them together
                component_key = ', '.join(sorted(components))
                by_component[component_key]['hours'] += seconds / 3600
                by_component[component_key]['entries'] += 1
                by_component[component_key]['issues'].add(issue_key)
            else:
                by_component['None']['hours'] += seconds / 3600
                by_component['None']['entries'] += 1
                by_component['None']['issues'].add(issue_key)

            # Aggregate by Label (keep multiple labels together, don't split)
            labels = worklog.get('labels', [])
            if labels:
                # Join multiple labels with comma to keep them together
                label_key = ', '.join(sorted(labels))
                by_label[label_key]['hours'] += seconds / 3600
                by_label[label_key]['entries'] += 1
                by_label[label_key]['issues'].add(issue_key)
            else:
                by_label['None']['hours'] += seconds / 3600
                by_label['None']['entries'] += 1
                by_label['None']['issues'].add(issue_key)

            # Aggregate by Team
            team = worklog.get('team', 'None')
            by_team[team]['hours'] += seconds / 3600
            by_team[team]['entries'] += 1
            by_team[team]['issues'].add(issue_key)

            total_seconds += seconds

        total_hours = total_seconds / 3600

        # Sort data
        authors_sorted = sorted(by_author.items(), key=lambda x: x[1]['hours'], reverse=True)
        issues_sorted = sorted(by_issue.items(), key=lambda x: x[1]['hours'], reverse=True)
        year_month_sorted = sorted(by_year_month.items(), key=lambda x: x[0])  # Sort by date
        product_items_sorted = sorted(by_product_item.items(), key=lambda x: x[1]['hours'], reverse=True)
        components_sorted = sorted(by_component.items(), key=lambda x: x[1]['hours'], reverse=True)
        labels_sorted = sorted(by_label.items(), key=lambda x: x[1]['hours'], reverse=True)
        teams_sorted = sorted(by_team.items(), key=lambda x: x[1]['hours'], reverse=True)

        # Generate HTML
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Jira Worklog Report - {self.project_key}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
            color: #333;
        }}

        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }}

        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}

        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
            font-weight: 600;
        }}

        .header .subtitle {{
            font-size: 1.1em;
            opacity: 0.9;
        }}

        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            padding: 40px;
            background: #f8f9fa;
        }}

        .stat-card {{
            background: white;
            padding: 25px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            text-align: center;
            transition: transform 0.2s;
        }}

        .stat-card:hover {{
            transform: translateY(-5px);
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }}

        .stat-card .number {{
            font-size: 2.5em;
            font-weight: bold;
            color: #667eea;
            margin-bottom: 5px;
        }}

        .stat-card .label {{
            color: #666;
            font-size: 0.9em;
            text-transform: uppercase;
            letter-spacing: 1px;
        }}

        .section {{
            padding: 40px;
        }}

        .section-title {{
            font-size: 1.8em;
            margin-bottom: 25px;
            color: #333;
            padding-bottom: 10px;
            border-bottom: 3px solid #667eea;
        }}

        .chart-container {{
            margin-bottom: 40px;
            background: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            background: white;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            border-radius: 8px;
            overflow: hidden;
        }}

        thead {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }}

        th {{
            padding: 15px;
            text-align: left;
            font-weight: 600;
            text-transform: uppercase;
            font-size: 0.85em;
            letter-spacing: 0.5px;
        }}

        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #eee;
        }}

        tbody tr:hover {{
            background: #f8f9fa;
        }}

        tbody tr:last-child td {{
            border-bottom: none;
        }}

        .progress-bar {{
            background: #e9ecef;
            height: 25px;
            border-radius: 12px;
            overflow: hidden;
            margin-top: 5px;
        }}

        .progress-fill {{
            background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
            height: 100%;
            transition: width 0.3s ease;
            display: flex;
            align-items: center;
            padding: 0 10px;
            color: white;
            font-size: 0.85em;
            font-weight: 600;
        }}

        .details-toggle {{
            background: #667eea;
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 0.9em;
            transition: background 0.2s;
        }}

        .details-toggle:hover {{
            background: #764ba2;
        }}

        .worklog-details {{
            display: none;
            margin-top: 15px;
            padding: 15px;
            background: #f8f9fa;
            border-radius: 8px;
            border-left: 4px solid #667eea;
        }}

        .worklog-details.active {{
            display: block;
        }}

        .worklog-entry {{
            background: white;
            padding: 12px;
            margin-bottom: 10px;
            border-radius: 4px;
            font-size: 0.9em;
        }}

        .worklog-entry:last-child {{
            margin-bottom: 0;
        }}

        .worklog-meta {{
            color: #666;
            font-size: 0.85em;
            margin-bottom: 5px;
        }}

        .worklog-comment {{
            color: #333;
            margin-top: 5px;
            font-style: italic;
        }}

        .footer {{
            background: #2d3748;
            color: white;
            padding: 20px 40px;
            text-align: center;
            font-size: 0.9em;
        }}

        .nav-menu {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px 40px;
            display: flex;
            flex-wrap: wrap;
            gap: 15px;
            justify-content: center;
            align-items: center;
        }}

        .nav-menu a {{
            color: white;
            text-decoration: none;
            padding: 10px 20px;
            background: rgba(255, 255, 255, 0.2);
            border-radius: 25px;
            font-size: 0.9em;
            font-weight: 500;
            transition: all 0.3s;
            backdrop-filter: blur(10px);
        }}

        .nav-menu a:hover {{
            background: rgba(255, 255, 255, 0.3);
            transform: translateY(-2px);
            box-shadow: 0 4px 8px rgba(0,0,0,0.2);
        }}

        @media print {{
            body {{
                background: white;
            }}
            .details-toggle {{
                display: none;
            }}
            .worklog-details {{
                display: block !important;
            }}
            .nav-menu {{
                display: none;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Jira Worklog Report</h1>
            <div class="subtitle">Project: {self.project_key}</div>
            <div class="subtitle">Version: {VERSION}</div>
            <div class="subtitle">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
        </div>

        <div class="stats-grid">
            <div class="stat-card">
                <div class="number">{total_hours:.1f}</div>
                <div class="label">Total Hours</div>
            </div>
            <div class="stat-card">
                <div class="number">{len(worklogs)}</div>
                <div class="label">Total Entries</div>
            </div>
            <div class="stat-card">
                <div class="number">{len(by_author)}</div>
                <div class="label">Contributors</div>
            </div>
            <div class="stat-card">
                <div class="number">{len(by_issue)}</div>
                <div class="label">Issues</div>
            </div>
        </div>

        <!-- Navigation Menu -->
        <div class="nav-menu">
            <a href="#insights">📊 Insights</a>
            <a href="#year-month">📅 By Year/Month</a>
            <a href="#product-item">📦 By Product Item</a>
            <a href="#component">🧩 By Component</a>
            <a href="#label">🏷️ By Label</a>
            <a href="#team">👥 By Team</a>
            <a href="#author">✍️ By Author</a>
            <a href="#issue">🎯 By Issue</a>
            <a href="#all-entries">📋 All Entries</a>
        </div>

        <!-- Analytics Section -->"""

        # Calculate advanced analytics
        # Most time-consuming issues
        issues_by_hours = sorted(by_issue.items(), key=lambda x: x[1]['hours'], reverse=True)
        top_issue = issues_by_hours[0] if issues_by_hours else None

        # Longest duration issues (time from first to last worklog)
        issue_duration = {}
        for issue_key in by_issue.keys():
            issue_worklogs = [w for w in worklogs if w['issue_key'] == issue_key]
            dates = [datetime.fromisoformat(w['started'].replace('Z', '+00:00')) for w in issue_worklogs if w['started']]
            if len(dates) > 1:
                duration_days = (max(dates) - min(dates)).days
                issue_duration[issue_key] = duration_days

        longest_issue = max(issue_duration.items(), key=lambda x: x[1]) if issue_duration else None

        # Average hours per contributor
        avg_hours_per_contributor = total_hours / len(by_author) if len(by_author) > 0 else 0

        # Most collaborative issue (most contributors)
        most_collaborative = max(by_issue.items(), key=lambda x: len(x[1]['authors'])) if by_issue else None

        # Issue type distribution
        type_distribution = defaultdict(lambda: {'count': 0, 'hours': 0})
        for issue_key, stats in by_issue.items():
            issue_type = stats['issue_type']
            type_distribution[issue_type]['count'] += 1
            type_distribution[issue_type]['hours'] += stats['hours']

        most_common_type = max(type_distribution.items(), key=lambda x: x[1]['count']) if type_distribution else None

        # Time period analysis
        all_dates = [datetime.fromisoformat(w['started'].replace('Z', '+00:00')) for w in worklogs if w['started']]
        if all_dates:
            first_date = min(all_dates)
            last_date = max(all_dates)
            total_days = (last_date - first_date).days + 1
            avg_hours_per_day = total_hours / total_days if total_days > 0 else 0
        else:
            first_date = last_date = None
            total_days = avg_hours_per_day = 0

        html += f"""
        <div class="section" id="insights">
            <h2 class="section-title">📊 Insights & Analytics</h2>
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 25px; border-radius: 12px; color: white; margin-bottom: 20px;">
                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px;">
                    <div style="background: rgba(255, 255, 255, 0.15); padding: 15px; border-radius: 8px; backdrop-filter: blur(10px);">
                        <div style="font-size: 0.9em; opacity: 0.9; margin-bottom: 5px;">⏱️ Time Period</div>
                        <div style="font-size: 1.3em; font-weight: bold;">{total_days} days</div>
                        <div style="font-size: 0.85em; opacity: 0.8; margin-top: 5px;">
                            {first_date.strftime('%Y-%m-%d') if first_date else 'N/A'} to {last_date.strftime('%Y-%m-%d') if last_date else 'N/A'}
                        </div>
                    </div>
                    <div style="background: rgba(255, 255, 255, 0.15); padding: 15px; border-radius: 8px; backdrop-filter: blur(10px);">
                        <div style="font-size: 0.9em; opacity: 0.9; margin-bottom: 5px;">📈 Avg. Hours/Day</div>
                        <div style="font-size: 1.3em; font-weight: bold;">{avg_hours_per_day:.1f}h</div>
                        <div style="font-size: 0.85em; opacity: 0.8; margin-top: 5px;">Active working days</div>
                    </div>
                    <div style="background: rgba(255, 255, 255, 0.15); padding: 15px; border-radius: 8px; backdrop-filter: blur(10px);">
                        <div style="font-size: 0.9em; opacity: 0.9; margin-bottom: 5px;">👥 Avg. Hours/Contributor</div>
                        <div style="font-size: 1.3em; font-weight: bold;">{avg_hours_per_contributor:.1f}h</div>
                        <div style="font-size: 0.85em; opacity: 0.8; margin-top: 5px;">{len(by_author)} total contributors</div>
                    </div>
                </div>
            </div>

            <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(350px, 1fr)); gap: 20px; margin-top: 20px;">
                <!-- Most Time Consuming Issue -->
                <div style="background: white; padding: 20px; border-radius: 8px; border-left: 4px solid #667eea; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                    <div style="font-size: 1.1em; font-weight: bold; color: #667eea; margin-bottom: 10px;">🔥 Most Time-Consuming Issue</div>
                    {f'<div style="font-weight: bold; margin-bottom: 5px;">{make_issue_link(top_issue[0])}</div>' if top_issue else '<div>N/A</div>'}
                    {f'<div style="color: #666; font-size: 0.9em; margin-bottom: 8px;">{top_issue[1]["summary"][:80]}...</div>' if top_issue else ''}
                    {f'<div style="font-size: 1.2em; color: #764ba2; font-weight: bold;">{top_issue[1]["hours"]:.1f} hours</div>' if top_issue else ''}
                    {f'<div style="font-size: 0.85em; color: #666; margin-top: 5px;">{top_issue[1]["entries"]} worklog entries • {len(top_issue[1]["authors"])} contributors</div>' if top_issue else ''}
                </div>

                <!-- Longest Duration Issue -->
                <div style="background: white; padding: 20px; border-radius: 8px; border-left: 4px solid #f59e0b; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                    <div style="font-size: 1.1em; font-weight: bold; color: #f59e0b; margin-bottom: 10px;">⏳ Longest Active Issue</div>
                    {f'<div style="font-weight: bold; margin-bottom: 5px;">{make_issue_link(longest_issue[0])}</div>' if longest_issue else '<div>N/A</div>'}
                    {f'<div style="color: #666; font-size: 0.9em; margin-bottom: 8px;">{by_issue[longest_issue[0]]["summary"][:80]}...</div>' if longest_issue else ''}
                    {f'<div style="font-size: 1.2em; color: #f59e0b; font-weight: bold;">{longest_issue[1]} days</div>' if longest_issue else ''}
                    {f'<div style="font-size: 0.85em; color: #666; margin-top: 5px;">{by_issue[longest_issue[0]]["hours"]:.1f} hours logged • Type: {by_issue[longest_issue[0]]["issue_type"]}</div>' if longest_issue else ''}
                </div>

                <!-- Most Collaborative Issue -->
                <div style="background: white; padding: 20px; border-radius: 8px; border-left: 4px solid #10b981; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                    <div style="font-size: 1.1em; font-weight: bold; color: #10b981; margin-bottom: 10px;">🤝 Most Collaborative Issue</div>
                    {f'<div style="font-weight: bold; margin-bottom: 5px;">{make_issue_link(most_collaborative[0])}</div>' if most_collaborative else '<div>N/A</div>'}
                    {f'<div style="color: #666; font-size: 0.9em; margin-bottom: 8px;">{most_collaborative[1]["summary"][:80]}...</div>' if most_collaborative else ''}
                    {f'<div style="font-size: 1.2em; color: #10b981; font-weight: bold;">{len(most_collaborative[1]["authors"])} contributors</div>' if most_collaborative else ''}
                    {f'<div style="font-size: 0.85em; color: #666; margin-top: 5px;">{most_collaborative[1]["hours"]:.1f} hours • {most_collaborative[1]["entries"]} entries</div>' if most_collaborative else ''}
                </div>

                <!-- Most Common Issue Type -->
                <div style="background: white; padding: 20px; border-radius: 8px; border-left: 4px solid #8b5cf6; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                    <div style="font-size: 1.1em; font-weight: bold; color: #8b5cf6; margin-bottom: 10px;">📋 Most Common Issue Type</div>
                    {f'<div style="font-size: 1.2em; color: #8b5cf6; font-weight: bold; margin-bottom: 5px;">{most_common_type[0]}</div>' if most_common_type else '<div>N/A</div>'}
                    {f'<div style="font-size: 0.9em; color: #666; margin-bottom: 8px;">{most_common_type[1]["count"]} issues of this type</div>' if most_common_type else ''}
                    {f'<div style="font-size: 1.1em; color: #333; font-weight: bold;">{most_common_type[1]["hours"]:.1f} hours</div>' if most_common_type else ''}
                    {f'<div style="font-size: 0.85em; color: #666; margin-top: 5px;">{(most_common_type[1]["count"] / len(by_issue) * 100):.1f}% of all issues</div>' if most_common_type else ''}
                </div>
            </div>

            <!-- Hours vs Estimate Analysis -->
            <div style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%); padding: 25px; border-radius: 12px; color: white; margin-top: 25px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                <div style="font-size: 1.3em; font-weight: bold; margin-bottom: 15px;">⚠️ Hours vs Original Estimate Analysis</div>
                <div style="background: rgba(255, 255, 255, 0.15); padding: 20px; border-radius: 8px; backdrop-filter: blur(10px);">
                    <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 20px;">
                        <div>
                            <div style="font-size: 0.9em; opacity: 0.9; margin-bottom: 5px;">Original Estimate</div>
                            <div style="font-size: 2em; font-weight: bold;">500h</div>
                        </div>
                        <div>
                            <div style="font-size: 0.9em; opacity: 0.9; margin-bottom: 5px;">Actual Hours Logged</div>
                            <div style="font-size: 2em; font-weight: bold;">{total_hours:.0f}h</div>
                        </div>
                        <div>
                            <div style="font-size: 0.9em; opacity: 0.9; margin-bottom: 5px;">Variance</div>
                            <div style="font-size: 2em; font-weight: bold;">+{(total_hours - 500):.0f}h</div>
                            <div style="font-size: 0.9em; opacity: 0.9; margin-top: 5px;">({((total_hours / 500 - 1) * 100):.0f}% over estimate)</div>
                        </div>
                    </div>
                    <div style="border-top: 1px solid rgba(255,255,255,0.3); padding-top: 15px; margin-top: 15px;">
                        <div style="font-size: 1.1em; font-weight: bold; margin-bottom: 10px;">🔍 Possible Reasons for Variance:</div>
                        <ul style="margin: 10px 0; padding-left: 20px; line-height: 1.8;">
                            <li style="margin-bottom: 8px;"><strong>Cross-project time logging:</strong> Hours from Unite project may have been logged here</li>
                            <li style="margin-bottom: 8px;"><strong>General development work:</strong> Development activities that should have been logged to IPP Development project</li>
                            <li style="margin-bottom: 8px;"><strong>Scope expansion:</strong> Additional requirements beyond original estimate (594 sub-tasks, {len(by_issue)} total issues)</li>
                            <li style="margin-bottom: 8px;"><strong>Underestimation:</strong> Original estimate may not have accounted for full project complexity</li>
                        </ul>
                        <div style="background: rgba(255,255,255,0.2); padding: 12px; border-radius: 6px; margin-top: 15px; font-size: 0.95em;">
                            💡 <strong>Recommendation:</strong> Review worklog entries to identify time logged from Unite project or general IPP development that should be reclassified to correct project codes.
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <div class="section" id="year-month">
            <h2 class="section-title">Hours by Year and Month</h2>

            <!-- Date Range Filter -->
            <div style="background: white; padding: 20px; border-radius: 8px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                <div style="display: flex; align-items: center; gap: 15px; flex-wrap: wrap;">
                    <label style="font-weight: 600; color: #333;">Filter by Date:</label>
                    <div style="display: flex; align-items: center; gap: 8px;">
                        <label for="dateFrom" style="color: #666;">From:</label>
                        <input type="month" id="dateFrom" style="padding: 8px; border: 1px solid #ddd; border-radius: 4px; font-size: 14px;">
                    </div>
                    <div style="display: flex; align-items: center; gap: 8px;">
                        <label for="dateTo" style="color: #666;">To:</label>
                        <input type="month" id="dateTo" style="padding: 8px; border: 1px solid #ddd; border-radius: 4px; font-size: 14px;">
                    </div>
                    <button onclick="applyDateFilter()" style="background: #667eea; color: white; border: none; padding: 10px 20px; border-radius: 4px; cursor: pointer; font-size: 14px; font-weight: 600; transition: background 0.2s;">
                        Apply Filter
                    </button>
                    <button onclick="resetDateFilter()" style="background: #6c757d; color: white; border: none; padding: 10px 20px; border-radius: 4px; cursor: pointer; font-size: 14px; font-weight: 600; transition: background 0.2s;">
                        Reset
                    </button>
                    <div id="filteredTotal" style="margin-left: auto; font-weight: 600; color: #667eea; display: none;">
                        Filtered Total: <span id="filteredHours">0</span>h
                    </div>
                </div>
            </div>

            <!-- Bar Chart -->
            <div class="chart-container">
                <canvas id="yearMonthChart" style="max-height: 400px;"></canvas>
            </div>

            <!-- Data Table -->
            <table>
                <thead>
                    <tr>
                        <th>Year-Month</th>
                        <th>Hours</th>
                        <th>Entries</th>
                        <th>% of Total</th>
                        <th>Distribution</th>
                    </tr>
                </thead>
                <tbody>
"""

        for year_month, stats in year_month_sorted:
            percentage = (stats['hours'] / total_hours * 100) if total_hours > 0 else 0
            html += f"""
                    <tr>
                        <td><strong>{year_month}</strong></td>
                        <td>{stats['hours']:.2f}h</td>
                        <td>{stats['entries']}</td>
                        <td>{percentage:.1f}%</td>
                        <td>
                            <div class="progress-bar">
                                <div class="progress-fill" style="width: {percentage}%">
                                    {percentage:.1f}%
                                </div>
                            </div>
                        </td>
                    </tr>
"""

        html += """
                </tbody>
            </table>
        </div>

        <div class="section" id="product-item">
            <h2 class="section-title">Hours by Product Item</h2>
            <table>
                <thead>
                    <tr>
                        <th>Product Item</th>
                        <th>Hours</th>
                        <th>Entries</th>
                        <th>Issues</th>
                        <th>% of Total</th>
                        <th>Distribution</th>
                    </tr>
                </thead>
                <tbody>
"""

        # Calculate totals for Product Item
        product_item_total_hours = sum(stats['hours'] for _, stats in product_items_sorted)
        product_item_total_entries = sum(stats['entries'] for _, stats in product_items_sorted)
        product_item_total_issues = len(set().union(*[stats['issues'] for _, stats in product_items_sorted]))

        for product_item, stats in product_items_sorted:
            percentage = (stats['hours'] / total_hours * 100) if total_hours > 0 else 0
            issue_count = len(stats['issues'])
            html += f"""
                    <tr>
                        <td>{product_item}</td>
                        <td>{stats['hours']:.2f}h</td>
                        <td>{stats['entries']}</td>
                        <td>{issue_count}</td>
                        <td>{percentage:.1f}%</td>
                        <td>
                            <div class="progress-bar">
                                <div class="progress-fill" style="width: {percentage}%">
                                    {percentage:.1f}%
                                </div>
                            </div>
                        </td>
                    </tr>
"""

        html += f"""
                </tbody>
                <tfoot>
                    <tr style="font-weight: bold; background-color: #f0f0f0;">
                        <td>TOTAL</td>
                        <td>{product_item_total_hours:.2f}h</td>
                        <td>{product_item_total_entries}</td>
                        <td>{product_item_total_issues}</td>
                        <td>100.0%</td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <div class="section" id="component">
            <h2 class="section-title">Hours by Component</h2>
            <table>
                <thead>
                    <tr>
                        <th>Component</th>
                        <th>Hours</th>
                        <th>Entries</th>
                        <th>Issues</th>
                        <th>% of Total</th>
                        <th>Distribution</th>
                    </tr>
                </thead>
                <tbody>
"""

        # Calculate totals for Component
        component_total_hours = sum(stats['hours'] for _, stats in components_sorted)
        component_total_entries = sum(stats['entries'] for _, stats in components_sorted)
        component_total_issues = len(set().union(*[stats['issues'] for _, stats in components_sorted]))

        for component, stats in components_sorted:
            percentage = (stats['hours'] / total_hours * 100) if total_hours > 0 else 0
            issue_count = len(stats['issues'])
            html += f"""
                    <tr>
                        <td>{component}</td>
                        <td>{stats['hours']:.2f}h</td>
                        <td>{stats['entries']}</td>
                        <td>{issue_count}</td>
                        <td>{percentage:.1f}%</td>
                        <td>
                            <div class="progress-bar">
                                <div class="progress-fill" style="width: {percentage}%">
                                    {percentage:.1f}%
                                </div>
                            </div>
                        </td>
                    </tr>
"""

        html += f"""
                </tbody>
                <tfoot>
                    <tr style="font-weight: bold; background-color: #f0f0f0;">
                        <td>TOTAL</td>
                        <td>{component_total_hours:.2f}h</td>
                        <td>{component_total_entries}</td>
                        <td>{component_total_issues}</td>
                        <td>100.0%</td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <div class="section" id="label">
            <h2 class="section-title">Hours by Label</h2>
            <table>
                <thead>
                    <tr>
                        <th>Label</th>
                        <th>Hours</th>
                        <th>Entries</th>
                        <th>Issues</th>
                        <th>% of Total</th>
                        <th>Distribution</th>
                    </tr>
                </thead>
                <tbody>
"""

        # Calculate totals for Label
        label_total_hours = sum(stats['hours'] for _, stats in labels_sorted)
        label_total_entries = sum(stats['entries'] for _, stats in labels_sorted)
        label_total_issues = len(set().union(*[stats['issues'] for _, stats in labels_sorted]))

        for label, stats in labels_sorted:
            percentage = (stats['hours'] / total_hours * 100) if total_hours > 0 else 0
            issue_count = len(stats['issues'])
            html += f"""
                    <tr>
                        <td>{label}</td>
                        <td>{stats['hours']:.2f}h</td>
                        <td>{stats['entries']}</td>
                        <td>{issue_count}</td>
                        <td>{percentage:.1f}%</td>
                        <td>
                            <div class="progress-bar">
                                <div class="progress-fill" style="width: {percentage}%">
                                    {percentage:.1f}%
                                </div>
                            </div>
                        </td>
                    </tr>
"""

        html += f"""
                </tbody>
                <tfoot>
                    <tr style="font-weight: bold; background-color: #f0f0f0;">
                        <td>TOTAL</td>
                        <td>{label_total_hours:.2f}h</td>
                        <td>{label_total_entries}</td>
                        <td>{label_total_issues}</td>
                        <td>100.0%</td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <div class="section" id="team">
            <h2 class="section-title">Hours by Team</h2>
            <table>
                <thead>
                    <tr>
                        <th>Team</th>
                        <th>Hours</th>
                        <th>Entries</th>
                        <th>Issues</th>
                        <th>% of Total</th>
                        <th>Distribution</th>
                    </tr>
                </thead>
                <tbody>
"""

        # Calculate totals for Team
        team_total_hours = sum(stats['hours'] for _, stats in teams_sorted)
        team_total_entries = sum(stats['entries'] for _, stats in teams_sorted)
        team_total_issues = len(set().union(*[stats['issues'] for _, stats in teams_sorted]))

        for team, stats in teams_sorted:
            percentage = (stats['hours'] / total_hours * 100) if total_hours > 0 else 0
            issue_count = len(stats['issues'])
            html += f"""
                    <tr>
                        <td>{team}</td>
                        <td>{stats['hours']:.2f}h</td>
                        <td>{stats['entries']}</td>
                        <td>{issue_count}</td>
                        <td>{percentage:.1f}%</td>
                        <td>
                            <div class="progress-bar">
                                <div class="progress-fill" style="width: {percentage}%">
                                    {percentage:.1f}%
                                </div>
                            </div>
                        </td>
                    </tr>
"""

        html += f"""
                </tbody>
                <tfoot>
                    <tr style="font-weight: bold; background-color: #f0f0f0;">
                        <td>TOTAL</td>
                        <td>{team_total_hours:.2f}h</td>
                        <td>{team_total_entries}</td>
                        <td>{team_total_issues}</td>
                        <td>100.0%</td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <div class="section" id="author">
            <h2 class="section-title">Hours by Author</h2>
            <table>
                <thead>
                    <tr>
                        <th>Author</th>
                        <th>Hours</th>
                        <th>Entries</th>
                        <th>% of Total</th>
                        <th>Distribution</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                <tbody>
"""

        # Calculate totals for Author
        author_total_hours = sum(stats['hours'] for _, stats in authors_sorted)
        author_total_entries = sum(stats['entries'] for _, stats in authors_sorted)

        for i, (author, stats) in enumerate(authors_sorted):
            percentage = (stats['hours'] / total_hours * 100) if total_hours > 0 else 0
            html += f"""
                    <tr>
                        <td><strong>{author}</strong></td>
                        <td>{stats['hours']:.2f}h</td>
                        <td>{stats['entries']}</td>
                        <td>{percentage:.1f}%</td>
                        <td>
                            <div class="progress-bar">
                                <div class="progress-fill" style="width: {percentage}%">
                                    {percentage:.1f}%
                                </div>
                            </div>
                        </td>
                        <td>
                            <button class="details-toggle" onclick="toggleDetails('author-{i}')">
                                View Details
                            </button>
                        </td>
                    </tr>
                    <tr>
                        <td colspan="6" style="padding: 0; border: none;">
                            <div id="author-{i}" class="worklog-details">
"""

            for worklog in sorted(stats['worklogs'], key=lambda x: x['started'], reverse=True):
                comment = worklog['comment'] if worklog['comment'] else 'No comment'
                epic_display = f" • Epic: {worklog['epic_link']}" if worklog['epic_link'] else ''
                issue_link = make_issue_link(worklog['issue_key'])
                html += f"""
                                <div class="worklog-entry">
                                    <div class="worklog-meta">
                                        {issue_link} ({worklog['issue_type']}) •
                                        {worklog['time_spent']} •
                                        {worklog['started'][:10]}{epic_display}
                                    </div>
                                    <div class="worklog-comment">{comment}</div>
                                </div>
"""

            html += """
                            </div>
                        </td>
                    </tr>
"""

        html += f"""
                </tbody>
                <tfoot>
                    <tr style="font-weight: bold; background-color: #f0f0f0;">
                        <td>TOTAL</td>
                        <td>{author_total_hours:.2f}h</td>
                        <td>{author_total_entries}</td>
                        <td>100.0%</td>
                        <td></td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <div class="section" id="issue">
            <h2 class="section-title">Hours by Issue</h2>
            <table>
                <thead>
                    <tr>
                        <th>Issue Key</th>
                        <th>Summary</th>
                        <th>Type</th>
                        <th>Epic Link</th>
                        <th>Product Item</th>
                        <th>Component</th>
                        <th>Label</th>
                        <th>Team</th>
                        <th>Hours</th>
                        <th>Entries</th>
                        <th>Contributors</th>
                        <th>% of Total</th>
                        <th>Distribution</th>
                    </tr>
                </thead>
                <tbody>
"""

        # Calculate totals for Issue
        issue_total_hours = sum(stats['hours'] for _, stats in issues_sorted)
        issue_total_entries = sum(stats['entries'] for _, stats in issues_sorted)
        issue_total_contributors = len(set().union(*[stats['authors'] for _, stats in issues_sorted]))

        for issue_key, stats in issues_sorted:
            percentage = (stats['hours'] / total_hours * 100) if total_hours > 0 else 0
            contributors = ', '.join(sorted(stats['authors']))
            epic_link = stats['epic_link'] if stats['epic_link'] else '-'
            summary = stats['summary'] if stats['summary'] else '-'

            # Format the 4 new fields
            components = ', '.join(stats.get('components', [])) if stats.get('components') else '-'
            labels = ', '.join(stats.get('labels', [])) if stats.get('labels') else '-'
            product_item = stats.get('product_item', 'None')
            team = stats.get('team', 'None')

            issue_link = make_issue_link(issue_key)
            html += f"""
                    <tr>
                        <td>{issue_link}</td>
                        <td>{summary}</td>
                        <td>{stats['issue_type']}</td>
                        <td>{epic_link}</td>
                        <td>{product_item}</td>
                        <td>{components}</td>
                        <td>{labels}</td>
                        <td>{team}</td>
                        <td>{stats['hours']:.2f}h</td>
                        <td>{stats['entries']}</td>
                        <td>{contributors}</td>
                        <td>{percentage:.1f}%</td>
                        <td>
                            <div class="progress-bar">
                                <div class="progress-fill" style="width: {percentage}%">
                                    {percentage:.1f}%
                                </div>
                            </div>
                        </td>
                    </tr>
"""

        html += f"""
                </tbody>
                <tfoot>
                    <tr style="font-weight: bold; background-color: #f0f0f0;">
                        <td colspan="8">TOTAL</td>
                        <td>{issue_total_hours:.2f}h</td>
                        <td>{issue_total_entries}</td>
                        <td>{issue_total_contributors} unique</td>
                        <td>100.0%</td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <div class="section" id="all-entries">
            <h2 class="section-title">All Worklog Entries</h2>
            <table>
                <thead>
                    <tr>
                        <th>Issue</th>
                        <th>Summary</th>
                        <th>Type</th>
                        <th>Epic Link</th>
                        <th>Author</th>
                        <th>Date</th>
                        <th>Time Spent</th>
                        <th>Comment</th>
                    </tr>
                </thead>
                <tbody>
"""

        for worklog in sorted(worklogs, key=lambda x: x['started'], reverse=True):
            comment = worklog['comment'][:100] + '...' if len(worklog['comment']) > 100 else worklog['comment']
            comment = comment if comment else 'No comment'
            epic_link = worklog['epic_link'] if worklog['epic_link'] else '-'
            summary = worklog['summary'][:50] + '...' if len(worklog['summary']) > 50 else worklog['summary']
            summary = summary if summary else '-'
            issue_link = make_issue_link(worklog['issue_key'])
            html += f"""
                    <tr>
                        <td>{issue_link}</td>
                        <td>{summary}</td>
                        <td>{worklog['issue_type']}</td>
                        <td>{epic_link}</td>
                        <td>{worklog['author']}</td>
                        <td>{worklog['started'][:10]}</td>
                        <td>{worklog['time_spent']}</td>
                        <td>{comment}</td>
                    </tr>
"""

        html += """
                </tbody>
            </table>
        </div>

        <div class="footer">
            Generated by Jira Worklog Extractor"""

        # Add timing information if available
        if timing_info:
            html += f"""<br>
            <div style="margin-top: 10px; font-size: 0.9em; opacity: 0.8;">
                Execution Time: {timing_info['total_time']:.2f}s
                (Collection: {timing_info['collection_time']:.2f}s,
                Extraction: {timing_info['extraction_time']:.2f}s) |
                {timing_info['issue_count']} issues,
                {timing_info['worklog_count']} worklogs
            </div>"""

        html += """
        </div>
    </div>

    <!-- Chart.js Library -->
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>

    <script>
        function toggleDetails(id) {
            const element = document.getElementById(id);
            element.classList.toggle('active');
        }

        // Year/Month Chart Data - Store original data
        const originalYearMonthLabels = """ + str([ym for ym, _ in year_month_sorted]) + """;
        const originalYearMonthHours = """ + str([stats['hours'] for _, stats in year_month_sorted]) + """;

        // Current chart instance
        let yearMonthChart = null;

        // Create the bar chart with initial data
        function createYearMonthChart(labels, hours) {
            const ctx = document.getElementById('yearMonthChart').getContext('2d');

            // Destroy existing chart if it exists
            if (yearMonthChart) {
                yearMonthChart.destroy();
            }

            yearMonthChart = new Chart(ctx, {
                type: 'bar',
                data: {
                    labels: labels,
                    datasets: [{
                        label: 'Hours Logged',
                        data: hours,
                        backgroundColor: 'rgba(102, 126, 234, 0.8)',
                        borderColor: 'rgba(102, 126, 234, 1)',
                        borderWidth: 2,
                        borderRadius: 5,
                        hoverBackgroundColor: 'rgba(118, 75, 162, 0.8)'
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: true,
                    plugins: {
                        legend: {
                            display: false
                        },
                        title: {
                            display: true,
                            text: 'Hours Logged by Month',
                            font: {
                                size: 16,
                                weight: 'bold'
                            },
                            padding: 20
                        },
                        tooltip: {
                            callbacks: {
                                label: function(context) {
                                    return 'Hours: ' + context.parsed.y.toFixed(2) + 'h';
                                }
                            }
                        }
                    },
                    scales: {
                        y: {
                            beginAtZero: true,
                            title: {
                                display: true,
                                text: 'Hours',
                                font: {
                                    size: 14,
                                    weight: 'bold'
                                }
                            },
                            ticks: {
                                callback: function(value) {
                                    return value.toFixed(0) + 'h';
                                }
                            },
                            grid: {
                                color: 'rgba(0, 0, 0, 0.05)'
                            }
                        },
                        x: {
                            title: {
                                display: true,
                                text: 'Year-Month',
                                font: {
                                    size: 14,
                                    weight: 'bold'
                                }
                            },
                            grid: {
                                display: false
                            }
                        }
                    }
                }
            });
        }

        // Date filter functions
        function applyDateFilter() {
            const fromInput = document.getElementById('dateFrom').value;
            const toInput = document.getElementById('dateTo').value;

            if (!fromInput && !toInput) {
                alert('Please select at least one date (from or to)');
                return;
            }

            // Filter the data
            const filteredLabels = [];
            const filteredHours = [];
            let totalFiltered = 0;

            for (let i = 0; i < originalYearMonthLabels.length; i++) {
                const label = originalYearMonthLabels[i];
                const hours = originalYearMonthHours[i];

                // Check if label is within range
                const inRange = (!fromInput || label >= fromInput) && (!toInput || label <= toInput);

                if (inRange) {
                    filteredLabels.push(label);
                    filteredHours.push(hours);
                    totalFiltered += hours;
                }
            }

            // Update chart with filtered data
            createYearMonthChart(filteredLabels, filteredHours);

            // Show filtered total
            document.getElementById('filteredHours').textContent = totalFiltered.toFixed(2);
            document.getElementById('filteredTotal').style.display = 'block';

            // Save to localStorage
            localStorage.setItem('dateFilterFrom', fromInput);
            localStorage.setItem('dateFilterTo', toInput);
        }

        function resetDateFilter() {
            // Clear inputs
            document.getElementById('dateFrom').value = '';
            document.getElementById('dateTo').value = '';

            // Reset chart to original data
            createYearMonthChart(originalYearMonthLabels, originalYearMonthHours);

            // Hide filtered total
            document.getElementById('filteredTotal').style.display = 'none';

            // Clear localStorage
            localStorage.removeItem('dateFilterFrom');
            localStorage.removeItem('dateFilterTo');
        }

        // Load date filter from localStorage on page load
        function loadDateFilter() {
            const fromValue = localStorage.getItem('dateFilterFrom');
            const toValue = localStorage.getItem('dateFilterTo');

            if (fromValue || toValue) {
                if (fromValue) document.getElementById('dateFrom').value = fromValue;
                if (toValue) document.getElementById('dateTo').value = toValue;
                applyDateFilter();
            } else {
                // Initialize chart with all data
                createYearMonthChart(originalYearMonthLabels, originalYearMonthHours);
            }
        }

        // Initialize chart on page load
        loadDateFilter();

        // =============================================================================
        // TABLE FILTERING AND SORTING FUNCTIONALITY
        // =============================================================================

        // Initialize filter state from localStorage
        let filterState = {
            globalSearch: '',
            productItem: '',
            component: '',
            label: '',
            team: '',
            author: ''
        };

        // Load filter state on page load
        function loadFilterState() {
            const saved = localStorage.getItem('jiraWorklogFilters');
            if (saved) {
                try {
                    filterState = JSON.parse(saved);
                } catch (e) {
                    console.error('Failed to parse saved filters:', e);
                }
            }
        }

        // Save filter state to localStorage
        function saveFilterState() {
            localStorage.setItem('jiraWorklogFilters', JSON.stringify(filterState));
        }

        // Clear all filters
        function clearAllFilters() {
            filterState = {
                globalSearch: '',
                productItem: '',
                component: '',
                label: '',
                team: '',
                author: ''
            };
            saveFilterState();

            // Clear all filter inputs
            document.querySelectorAll('.global-filter').forEach(input => input.value = '');
            document.querySelectorAll('.column-filter').forEach(select => select.value = '');

            // Reset all table rows
            document.querySelectorAll('table tbody tr').forEach(row => {
                row.style.display = '';
            });

            // Update visible counts
            updateVisibleCounts();
        }

        // Apply all active filters to a table
        function applyFilters(tableId) {
            const table = document.getElementById(tableId);
            if (!table) return;

            const tbody = table.querySelector('tbody');
            if (!tbody) return;

            const rows = tbody.querySelectorAll('tr:not(.worklog-details-row)');
            let visibleCount = 0;

            rows.forEach(row => {
                // Skip detail rows
                if (row.querySelector('.worklog-details')) {
                    return;
                }

                let visible = true;

                // Apply global search filter
                if (filterState.globalSearch) {
                    const searchTerm = filterState.globalSearch.toLowerCase();
                    const rowText = row.textContent.toLowerCase();
                    if (!rowText.includes(searchTerm)) {
                        visible = false;
                    }
                }

                // Apply column-specific filters
                const cells = row.querySelectorAll('td');

                // Product Item filter (if column exists)
                if (filterState.productItem && tableId === 'issueTable') {
                    const productItemCell = cells[4]; // Column index for Product Item
                    if (productItemCell && productItemCell.textContent.trim() !== filterState.productItem) {
                        visible = false;
                    }
                }

                // Component filter (if column exists)
                if (filterState.component && tableId === 'issueTable') {
                    const componentCell = cells[5]; // Column index for Component
                    if (componentCell) {
                        const components = componentCell.textContent.trim();
                        if (!components.includes(filterState.component) && filterState.component !== '-') {
                            visible = false;
                        } else if (filterState.component === '-' && components !== '-') {
                            visible = false;
                        }
                    }
                }

                // Label filter (if column exists)
                if (filterState.label && tableId === 'issueTable') {
                    const labelCell = cells[6]; // Column index for Label
                    if (labelCell) {
                        const labels = labelCell.textContent.trim();
                        if (!labels.includes(filterState.label) && filterState.label !== '-') {
                            visible = false;
                        } else if (filterState.label === '-' && labels !== '-') {
                            visible = false;
                        }
                    }
                }

                // Team filter (if column exists)
                if (filterState.team && tableId === 'issueTable') {
                    const teamCell = cells[7]; // Column index for Team
                    if (teamCell && teamCell.textContent.trim() !== filterState.team) {
                        visible = false;
                    }
                }

                // Author filter (if column exists)
                if (filterState.author && (tableId === 'authorTable' || tableId === 'allEntriesTable')) {
                    let authorCell;
                    if (tableId === 'authorTable') {
                        authorCell = cells[0]; // First column in author table
                    } else if (tableId === 'allEntriesTable') {
                        authorCell = cells[4]; // Author column in all entries
                    }
                    if (authorCell && authorCell.textContent.trim() !== filterState.author) {
                        visible = false;
                    }
                }

                row.style.display = visible ? '' : 'none';
                if (visible) visibleCount++;
            });

            return visibleCount;
        }

        // Update visible row counts
        function updateVisibleCounts() {
            const tables = ['issueTable', 'authorTable', 'allEntriesTable', 'productItemTable',
                          'componentTable', 'labelTable', 'teamTable'];

            tables.forEach(tableId => {
                const count = applyFilters(tableId);
                const countElement = document.getElementById(tableId + 'Count');
                if (countElement) {
                    countElement.textContent = count;
                }
            });
        }

        // Create filter controls for a table
        function createFilterControls(tableId, columnFilters = []) {
            const table = document.getElementById(tableId);
            if (!table) return;

            const section = table.closest('.section');
            if (!section) return;

            // Create filter container
            const filterContainer = document.createElement('div');
            filterContainer.className = 'filter-container';
            filterContainer.style.cssText = `
                background: #f8f9fa;
                padding: 20px;
                border-radius: 8px;
                margin-bottom: 20px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            `;

            // Global search
            const searchDiv = document.createElement('div');
            searchDiv.style.marginBottom = '15px';
            searchDiv.innerHTML = `
                <label style="display: block; font-weight: 600; margin-bottom: 8px; color: #333;">
                    🔍 Search:
                </label>
                <div style="display: flex; gap: 10px;">
                    <input type="text"
                           class="global-filter"
                           data-table="${tableId}"
                           placeholder="Filter all columns..."
                           style="flex: 1; padding: 10px; border: 1px solid #ddd; border-radius: 4px; font-size: 14px;">
                    <button onclick="clearSearchFor('${tableId}')"
                            style="padding: 10px 20px; background: #6c757d; color: white; border: none;
                                   border-radius: 4px; cursor: pointer; font-size: 14px;">
                        Clear
                    </button>
                </div>
            `;
            filterContainer.appendChild(searchDiv);

            // Column-specific filters
            if (columnFilters.length > 0) {
                const columnFiltersDiv = document.createElement('div');
                columnFiltersDiv.style.cssText = `
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                    gap: 15px;
                `;

                columnFilters.forEach(filter => {
                    const filterDiv = document.createElement('div');
                    filterDiv.innerHTML = `
                        <label style="display: block; font-weight: 600; margin-bottom: 8px; color: #333;">
                            ${filter.icon} ${filter.label}:
                        </label>
                        <select class="column-filter"
                                data-filter="${filter.key}"
                                data-table="${tableId}"
                                style="width: 100%; padding: 10px; border: 1px solid #ddd;
                                       border-radius: 4px; font-size: 14px; background: white;">
                            <option value="">All ${filter.label}</option>
                            ${filter.options.map(opt =>
                                `<option value="${opt}">${opt}</option>`
                            ).join('')}
                        </select>
                    `;
                    columnFiltersDiv.appendChild(filterDiv);
                });

                filterContainer.appendChild(columnFiltersDiv);
            }

            // Visible count display
            const countDiv = document.createElement('div');
            countDiv.style.cssText = 'margin-top: 15px; font-weight: 600; color: #667eea;';
            countDiv.innerHTML = `
                Showing <span id="${tableId}Count">0</span> rows
            `;
            filterContainer.appendChild(countDiv);

            // Insert before table
            table.parentNode.insertBefore(filterContainer, table);

            // Restore saved filters
            const globalInput = filterContainer.querySelector('.global-filter');
            if (globalInput) {
                globalInput.value = filterState.globalSearch;
            }

            filterContainer.querySelectorAll('.column-filter').forEach(select => {
                const filterKey = select.dataset.filter;
                if (filterState[filterKey]) {
                    select.value = filterState[filterKey];
                }
            });
        }

        // Clear search for specific table
        function clearSearchFor(tableId) {
            const input = document.querySelector(`.global-filter[data-table="${tableId}"]`);
            if (input) {
                input.value = '';
                filterState.globalSearch = '';
                saveFilterState();
                applyFilters(tableId);
                updateVisibleCounts();
            }
        }

        // Extract unique values from table column
        function getUniqueColumnValues(tableId, columnIndex) {
            const table = document.getElementById(tableId);
            if (!table) return [];

            const values = new Set();
            const rows = table.querySelectorAll('tbody tr:not(.worklog-details-row)');

            rows.forEach(row => {
                if (row.querySelector('.worklog-details')) return;

                const cells = row.querySelectorAll('td');
                if (cells[columnIndex]) {
                    const text = cells[columnIndex].textContent.trim();
                    if (text) {
                        // For multi-value cells (like components), split and add individually
                        if (text.includes(',')) {
                            text.split(',').forEach(val => {
                                const trimmed = val.trim();
                                if (trimmed) values.add(trimmed);
                            });
                        } else {
                            values.add(text);
                        }
                    }
                }
            });

            return Array.from(values).sort();
        }

        // =============================================================================
        // TABLE SORTING FUNCTIONALITY
        // =============================================================================

        let sortState = {};

        function makeSortable(tableId) {
            const table = document.getElementById(tableId);
            if (!table) return;

            const headers = table.querySelectorAll('thead th');

            headers.forEach((header, index) => {
                // Skip action columns
                if (header.textContent.includes('Actions') || header.textContent.includes('Distribution')) {
                    return;
                }

                header.style.cursor = 'pointer';
                header.style.userSelect = 'none';
                header.style.position = 'relative';
                header.style.paddingRight = '25px';

                // Add sort indicator
                const indicator = document.createElement('span');
                indicator.className = 'sort-indicator';
                indicator.style.cssText = `
                    position: absolute;
                    right: 8px;
                    top: 50%;
                    transform: translateY(-50%);
                    opacity: 0.3;
                `;
                indicator.innerHTML = '▲▼';
                header.appendChild(indicator);

                header.addEventListener('click', (e) => {
                    if (e.shiftKey) {
                        // Multi-column sort (future enhancement)
                        sortTable(tableId, index, true);
                    } else {
                        sortTable(tableId, index, false);
                    }
                });
            });
        }

        function sortTable(tableId, columnIndex, multiSort = false) {
            const table = document.getElementById(tableId);
            if (!table) return;

            const tbody = table.querySelector('tbody');
            const rows = Array.from(tbody.querySelectorAll('tr:not(.worklog-details-row)'));

            // Filter out detail rows
            const dataRows = rows.filter(row => !row.querySelector('.worklog-details'));

            // Determine sort direction
            const currentSort = sortState[tableId] || {};
            const currentDirection = currentSort.column === columnIndex ? currentSort.direction : null;
            const newDirection = currentDirection === 'asc' ? 'desc' : 'asc';

            // Update sort state
            sortState[tableId] = { column: columnIndex, direction: newDirection };

            // Update sort indicators
            table.querySelectorAll('thead th .sort-indicator').forEach((indicator, idx) => {
                if (idx === columnIndex) {
                    indicator.style.opacity = '1';
                    indicator.innerHTML = newDirection === 'asc' ? '▲' : '▼';
                } else {
                    indicator.style.opacity = '0.3';
                    indicator.innerHTML = '▲▼';
                }
            });

            // Sort rows
            dataRows.sort((a, b) => {
                const cellA = a.querySelectorAll('td')[columnIndex];
                const cellB = b.querySelectorAll('td')[columnIndex];

                if (!cellA || !cellB) return 0;

                let valueA = cellA.textContent.trim();
                let valueB = cellB.textContent.trim();

                // Try to parse as number
                const numA = parseFloat(valueA.replace(/[^0-9.-]/g, ''));
                const numB = parseFloat(valueB.replace(/[^0-9.-]/g, ''));

                if (!isNaN(numA) && !isNaN(numB)) {
                    return newDirection === 'asc' ? numA - numB : numB - numA;
                }

                // String comparison
                return newDirection === 'asc'
                    ? valueA.localeCompare(valueB)
                    : valueB.localeCompare(valueA);
            });

            // Re-append sorted rows
            dataRows.forEach(row => {
                tbody.appendChild(row);
                // If there's a detail row, append it after
                const detailRow = row.nextElementSibling;
                if (detailRow && detailRow.querySelector('.worklog-details')) {
                    tbody.appendChild(detailRow);
                }
            });
        }

        // =============================================================================
        // INITIALIZATION
        // =============================================================================

        document.addEventListener('DOMContentLoaded', function() {
            // Load saved filter state
            loadFilterState();

            // Create tables with IDs for easier targeting
            const issueTable = document.querySelector('#issue table');
            if (issueTable) issueTable.id = 'issueTable';

            const authorTable = document.querySelector('#author table');
            if (authorTable) authorTable.id = 'authorTable';

            const allEntriesTable = document.querySelector('#all-entries table');
            if (allEntriesTable) allEntriesTable.id = 'allEntriesTable';

            const productItemTable = document.querySelector('#product-item table');
            if (productItemTable) productItemTable.id = 'productItemTable';

            const componentTable = document.querySelector('#component table');
            if (componentTable) componentTable.id = 'componentTable';

            const labelTable = document.querySelector('#label table');
            if (labelTable) labelTable.id = 'labelTable';

            const teamTable = document.querySelector('#team table');
            if (teamTable) teamTable.id = 'teamTable';

            // Setup filters for Issue table
            if (issueTable) {
                const productItems = getUniqueColumnValues('issueTable', 4);
                const components = getUniqueColumnValues('issueTable', 5);
                const labels = getUniqueColumnValues('issueTable', 6);
                const teams = getUniqueColumnValues('issueTable', 7);

                createFilterControls('issueTable', [
                    { key: 'productItem', label: 'Product Item', icon: '📦', options: productItems },
                    { key: 'component', label: 'Component', icon: '🧩', options: components },
                    { key: 'label', label: 'Label', icon: '🏷️', options: labels },
                    { key: 'team', label: 'Team', icon: '👥', options: teams }
                ]);
                makeSortable('issueTable');
            }

            // Setup filters for Author table
            if (authorTable) {
                const authors = getUniqueColumnValues('authorTable', 0);
                createFilterControls('authorTable', [
                    { key: 'author', label: 'Author', icon: '✍️', options: authors }
                ]);
                makeSortable('authorTable');
            }

            // Setup filters for All Entries table
            if (allEntriesTable) {
                createFilterControls('allEntriesTable', []);
                makeSortable('allEntriesTable');
            }

            // Setup filters for aggregate tables
            if (productItemTable) {
                createFilterControls('productItemTable', []);
                makeSortable('productItemTable');
            }

            if (componentTable) {
                createFilterControls('componentTable', []);
                makeSortable('componentTable');
            }

            if (labelTable) {
                createFilterControls('labelTable', []);
                makeSortable('labelTable');
            }

            if (teamTable) {
                createFilterControls('teamTable', []);
                makeSortable('teamTable');
            }

            // Attach event listeners to all filter inputs
            document.querySelectorAll('.global-filter').forEach(input => {
                input.addEventListener('input', function() {
                    filterState.globalSearch = this.value;
                    saveFilterState();

                    const tableId = this.dataset.table;
                    applyFilters(tableId);
                    updateVisibleCounts();
                });
            });

            document.querySelectorAll('.column-filter').forEach(select => {
                select.addEventListener('change', function() {
                    const filterKey = this.dataset.filter;
                    filterState[filterKey] = this.value;
                    saveFilterState();

                    const tableId = this.dataset.table;
                    applyFilters(tableId);
                    updateVisibleCounts();
                });
            });

            // Add "Clear All Filters" button to navigation
            const navMenu = document.querySelector('.nav-menu');
            if (navMenu) {
                const clearButton = document.createElement('a');
                clearButton.href = '#';
                clearButton.innerHTML = '🗑️ Clear All Filters';
                clearButton.style.background = 'rgba(239, 68, 68, 0.8)';
                clearButton.addEventListener('click', function(e) {
                    e.preventDefault();
                    clearAllFilters();
                });
                navMenu.appendChild(clearButton);
            }

            // Apply initial filters (restore from localStorage)
            updateVisibleCounts();
        });
    </script>
</body>
</html>
"""

        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html)

        print(f"HTML report exported to {filename}")


def parse_arguments():
    """Parse command-line arguments"""
    parser = argparse.ArgumentParser(
        description=f'Jira Worklog Extractor v{VERSION} - Extract worklogs from Jira filtered by ERP Activity',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  # Use environment variables from .env file
  python extract_worklogs.py

  # Override with command-line arguments
  python extract_worklogs.py --project ZYN --erp-activity ProjectTask-00000007118797

  # Debug mode with custom output directory
  python extract_worklogs.py --log-level 2 --output-dir ./reports

  # Export only CSV
  python extract_worklogs.py --format csv
        '''
    )

    parser.add_argument('--project', '-p',
                        default=PROJECT_KEY,
                        help=f'Project key (default: {PROJECT_KEY})')

    parser.add_argument('--erp-activity', '-e',
                        default=ERP_ACTIVITY_FILTER,
                        help=f'ERP Activity filter value (default: {ERP_ACTIVITY_FILTER})')

    parser.add_argument('--jira-url',
                        default=JIRA_URL,
                        help=f'Jira instance URL (default: from env or {JIRA_URL})')

    parser.add_argument('--jira-token',
                        default=JIRA_API_TOKEN,
                        help='Jira API token (default: from env)')

    parser.add_argument('--output-dir', '-o',
                        default='output',
                        help='Output directory for reports (default: ./output)')

    parser.add_argument('--log-level', '-l',
                        type=int,
                        choices=[1, 2],
                        default=LOG_LEVEL,
                        help='Log level: 1=standard, 2=debug (default: 1)')

    parser.add_argument('--format', '-f',
                        choices=['csv', 'html', 'excel', 'both', 'all'],
                        default='both',
                        help='Output format: csv, html, excel, both (csv+html), or all (csv+html+excel) (default: both)')

    parser.add_argument('--version', '-v',
                        action='version',
                        version=f'%(prog)s {VERSION}')

    parser.add_argument('--no-cache',
                        action='store_true',
                        help='Disable caching of API responses')

    parser.add_argument('--cache-ttl',
                        type=int,
                        default=3600,
                        help='Cache time-to-live in seconds (default: 3600 = 1 hour)')

    parser.add_argument('--skip-validation',
                        action='store_true',
                        help='Skip configuration validation on startup')

    return parser.parse_args()


def main():
    # Parse command-line arguments
    args = parse_arguments()

    # Setup logging
    logger = setup_logging(log_level=args.log_level)

    # Display configuration
    print(f"\n{'='*70}")
    print(f"Jira Worklog Extractor v{VERSION}")
    print(f"{'='*70}")
    print(f"Configuration:")
    print(f"  Jira URL: {args.jira_url}")
    print(f"  Project: {args.project}")
    print(f"  ERP Activity: {args.erp_activity}")
    print(f"  Output Directory: {args.output_dir}")
    print(f"  Log Level: {args.log_level}")
    print(f"  Format: {args.format}")
    print(f"  Cache: {'Disabled' if args.no_cache else f'Enabled (TTL: {args.cache_ttl}s)'}")
    print(f"{'='*70}\n")

    # Track execution time
    start_time = time.time()

    # Initialize extractor
    extractor = JiraWorklogExtractor(
        jira_url=args.jira_url,
        api_token=args.jira_token,
        project_key=args.project,
        erp_activity_filter=args.erp_activity,
        log_level=args.log_level,
        use_cache=not args.no_cache,
        cache_ttl=args.cache_ttl,
        skip_validation=args.skip_validation,
        logger=logger
    )

    # Collect all related issues
    collection_start = time.time()
    issue_keys = extractor.collect_all_related_issues()
    collection_time = time.time() - collection_start

    if not issue_keys:
        print("No issues found matching the criteria")
        return

    # Extract worklogs
    extraction_start = time.time()
    worklogs = extractor.extract_worklogs(issue_keys)
    extraction_time = time.time() - extraction_start

    # Generate summary
    extractor.generate_summary(worklogs)

    # Create output directory if it doesn't exist
    output_dir = Path(args.output_dir)
    output_dir.mkdir(exist_ok=True)

    # Generate filename without timestamp (will overwrite existing files)
    base_filename = f'{args.project.lower()}_worklogs'

    # Export based on format selection
    print("\nExporting results...")

    if args.format in ['csv', 'both', 'all']:
        csv_filename = output_dir / f'{base_filename}.csv'
        extractor.export_to_csv(worklogs, str(csv_filename))
        print(f"  - CSV: {csv_filename}")

    if args.format in ['html', 'both', 'all']:
        html_filename = output_dir / f'{base_filename}.html'
        total_time = time.time() - start_time
        timing_info = {
            'collection_time': collection_time,
            'extraction_time': extraction_time,
            'total_time': total_time,
            'issue_count': len(issue_keys),
            'worklog_count': len(worklogs)
        }
        extractor.export_to_html(worklogs, str(html_filename), timing_info)
        print(f"  - HTML Report: {html_filename}")

    if args.format in ['excel', 'all']:
        excel_filename = output_dir / f'{base_filename}.xlsx'
        extractor.export_to_excel(worklogs, str(excel_filename))
        print(f"  - Excel: {excel_filename}")

    total_execution_time = time.time() - start_time
    print(f"\nDone! Results exported to {output_dir}")
    print(f"Total execution time: {total_execution_time:.2f}s")


if __name__ == "__main__":
    main()
